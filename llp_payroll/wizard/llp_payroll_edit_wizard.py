# -*- coding: utf-8 -*-
from odoo import api, fields, models, _ # type: ignore
from odoo.exceptions import UserError # type: ignore
from io import BytesIO
import base64
from collections import defaultdict

try:
    import openpyxl # type: ignore
except ImportError:
    openpyxl = None

import xlsxwriter # type: ignore


class LLPPayrollEditWizard(models.TransientModel):
    _name = 'llp.payroll.edit.wizard'
    _description = 'Payroll Edit Wizard (Export/Import)'

    mode = fields.Selection([('export', 'Export'), ('import', 'Import')], default='export', required=True)
    payroll_id = fields.Many2one('llp.payroll', string="Payroll", required=True)
    file_data = fields.Binary(string="Excel File")
    file_name = fields.Char(string="File Name")

    # -----------------------------
    # EXPORT (matrix)
    # -----------------------------
    def action_export(self):
        self.ensure_one()

        query = """
            SELECT
                G.id as employee_id,
                G.lastname,
                G.firstname,
                G.identification_id,
                D.id as rule_id,
                D.name as rule_name,
                D.code as rule_code,
                COALESCE(F.value, 0) as value,
                C.sequence as seq
            FROM llp_payroll A
                LEFT JOIN llp_payroll_structure B ON B.id = A.struct_id
                LEFT JOIN llp_payroll_structure_line C ON C.struct_id = B.id
                INNER JOIN llp_payroll_rule D ON D.id = C.rule_id AND D.ruleview_type = 'edit'
                LEFT JOIN llp_payroll_line E ON E.payroll_id = A.id
                LEFT JOIN llp_payroll_rule_value F ON F.payroll_rule_id = D.id AND F.line_id = E.id
                LEFT JOIN hr_employee G ON G.id = E.employee_id
            WHERE A.id = %s
            ORDER BY G.lastname, G.firstname, C.sequence
        """
        self.env.cr.execute(query, (self.payroll_id.id,))
        rows = self.env.cr.fetchall()
        if not rows:
            raise UserError(_("Засах боломжтой өгөгдөл олдсонгүй."))

        employees = {}
        rules = []          # list of (rule_id, name, code) in order
        rule_index = {}     # rule_id -> index

        for emp_id, last, first, identification, rule_id, rule_name, rule_code, value, seq in rows:
            if rule_id not in rule_index:
                rule_index[rule_id] = len(rules)
                rules.append((rule_id, rule_name or '', rule_code or ''))

            if emp_id not in employees:
                employees[emp_id] = {
                    'last': last or '',
                    'first': first or '',
                    'identification': identification or '',
                    'values': defaultdict(float)
                }

            employees[emp_id]['values'][rule_id] += float(value or 0.0)

        # ---- Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet("Payroll")

        header_fmt = workbook.add_format({'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter'})
        text_fmt = workbook.add_format({'border': 1})
        num_fmt = workbook.add_format({'border': 1, 'align': 'right', 'num_format': '#,##0.00'})
        total_lbl_fmt = workbook.add_format({'border': 1, 'bold': True, 'align': 'right'})
        total_num_fmt = workbook.add_format({'border': 1, 'bold': True, 'align': 'right', 'num_format': '#,##0.00'})

        # Fixed cols
        sheet.set_column(0, 0, 16)
        sheet.set_column(1, 1, 16)
        sheet.set_column(2, 2, 18)

        sheet.merge_range(0, 0, 1, 0, "Овог", header_fmt)
        sheet.merge_range(0, 1, 1, 1, "Нэр", header_fmt)
        sheet.merge_range(0, 2, 1, 2, "Регистрийн дугаар", header_fmt)

        start_rule_col = 3
        for i, (rid, rname, rcode) in enumerate(rules):
            col = start_rule_col + i
            sheet.write(0, col, rname, header_fmt)
            sheet.write(1, col, rcode, header_fmt)
            sheet.set_column(col, col, 18)

        # Data
        row = 2
        column_totals = [0.0] * len(rules)

        for emp in sorted(employees.values(), key=lambda e: (e['last'], e['first'])):
            sheet.write(row, 0, emp['last'], text_fmt)
            sheet.write(row, 1, emp['first'], text_fmt)
            sheet.write(row, 2, emp['identification'], text_fmt)

            for rule_id, idx in rule_index.items():
                val = emp['values'].get(rule_id, 0.0)
                sheet.write_number(row, start_rule_col + idx, val, num_fmt)
                column_totals[idx] += val

            row += 1

        workbook.close()

        file_name = "Цалин бодолт.xlsx"
        data = base64.b64encode(output.getvalue())
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'type': 'binary',
            'datas': data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{attachment.id}?download=true', 'target': 'self'}

    def action_import(self):
        self.ensure_one()

        if not openpyxl:
            raise UserError(_("openpyxl is not installed on server."))

        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))

        content = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active

        header_row = 2
        data_start_row = 3

        ident_col = 3
        first_rule_col = 4

        col_rule_code = {}
        max_col = ws.max_column

        for col in range(first_rule_col, max_col + 1):
            code = ws.cell(row=header_row, column=col).value
            if code:
                col_rule_code[col] = str(code).strip()

        if not col_rule_code:
            raise UserError(_("No rule codes found in header row."))

        rule_codes = list(set(col_rule_code.values()))
        rules = self.env['llp.payroll.rule'].sudo().search([('code', 'in', rule_codes)])
        rule_by_code = {r.code: r for r in rules}

        missing = [c for c in rule_codes if c not in rule_by_code]
        if missing:
            raise UserError(_("These rule codes do not exist in llp.payroll.rule:\n%s") % ", ".join(missing))

        plines = self.env['llp.payroll.line'].sudo().search([('payroll_id', '=', self.payroll_id.id)])
        line_by_emp = {l.employee_id.id: l for l in plines if l.employee_id}

        RV = self.env['llp.payroll.rule.value'].sudo()

        updates = 0

        for r in range(data_start_row, ws.max_row + 1):
            ident = ws.cell(row=r, column=ident_col).value
            if not ident:
                continue
            ident = str(ident).strip()

            emp = self.env['hr.employee'].sudo().search([('identification_id', '=', ident)], limit=1)
            if not emp:
                continue

            pline = line_by_emp.get(emp.id)
            if not pline:
                continue

            for col, code in col_rule_code.items():
                cell_val = ws.cell(row=r, column=col).value
                if cell_val is None or cell_val == '':
                    continue

                try:
                    val = float(cell_val)
                except Exception:
                    continue

                rule = rule_by_code[code]

                rv = RV.search([('payroll_rule_id', '=', rule.id), ('line_id', '=', pline.id)], limit=1)
                if rv:
                    rv.write({'value': val})
                    updates += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'reload'
        }